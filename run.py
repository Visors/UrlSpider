from BaiduSpider import BaiduSpider
import sys
from openpyxl import Workbook, load_workbook

out_xlsx = Workbook()
baidu = out_xlsx.active
baidu.title = '百度'


def main(name, company, keyword):
    if keyword is None:
        keyword = company + name
    baidu_spider = BaiduSpider()
    data = baidu_spider.search(keyword)
    while not data.empty():
        item = data.get()
        if name in item.title or company in item.title:
            baidu.append([name, company, item.title, item.link, 'Y'])
        else:
            baidu.append([name, company, item.title, item.link, 'N'])


if __name__ == '__main__':
    in_xlsx = load_workbook('查询表单.xlsx')
    in_sheet = in_xlsx.active
    for row in in_sheet.iter_rows():
        name = row[0].value  # 检索人名
        company = row[1].value  # 检索人所在机构
        keyword = row[2].value  # 检索关键词
        print(name, company, keyword)
        main(name, company, keyword)
        out_xlsx.save('查询结果.xlsx')
    out_xlsx.save('查询结果.xlsx')
    print('Done!')

# if len(sys.argv) != 2:
#     print('no keyword!')
#     sys.exit(-1)
# else:
#     main(sys.argv[1])
